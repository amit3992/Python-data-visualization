__author__ = 'kulkara0'

import openpyxl
import pprint

wb1 = openpyxl.load_workbook('0909_WN9.1Baselines_Report.xlsx')

activeSheet1 = wb1.active
maxRow = activeSheet1.max_row

activeSheet1.columns[0]
data = []           # Store row numbers
scripts =[]
scriptData = {}

for cellObj in activeSheet1.columns[0]:
    if(cellObj.value == "Script Name"):
         data.append(cellObj.row)
x = data[0]


for k in range(x,maxRow,1):
    val = activeSheet1.cell(row=k,column=1).value
    if(type(val) == str and val != "Script Name"):
        scriptName  = activeSheet1['A' + str(k)].value
        transName = activeSheet1['B' + str(k)].value

         # Make sure the key for this state exists.
        scriptData.setdefault(scriptName, {})
        # Make sure the key for this county in this state exists
        scriptData[scriptName].setdefault(transName, {'Avg': 0, 'Max': 0, 'nine': 0, 'nine5': 0})

        # Each row represents one census tract, so increment by one.
        scriptData[scriptName][transName]['Avg'] = float(activeSheet1.cell(row=k,column=6).value)
        #print(activeSheet1.cell(row=k,column=6).value)
        #print(activeSheet1['F' + str(k)].value)

        scriptData[scriptName][transName]['Max'] = float(activeSheet1['H' + str(k)].value)
        #print(activeSheet1.cell(row=k,column=8).value)
        #print(activeSheet1['H' + str(k)].value)

        scriptData[scriptName][transName]['nine'] = float(activeSheet1['I' + str(k)].value)
        #print(activeSheet1.cell(row=k,column=9).value)
        #print(activeSheet1['I' + str(k)].value)

        scriptData[scriptName][transName]['nine5'] = float(activeSheet1['J' + str(k)].value)
        #print(activeSheet1.cell(row=k,column=10).value)
       #print(activeSheet1['J' + str(k)].value)

print('Writing results...')
resultFile = open('scriptDictionary.py', 'w')
resultFile.write('allData = ' + pprint.pformat(scriptData))
resultFile.close()
print('Done.')
