__author__ = 'kulkara0'

import openpyxl
import pprint
import matplotlib
import xlsxwriter

# ============================================= Baseline Reports ===============================================

wb1 = openpyxl.load_workbook('0909_WN9.1Baselines_Report.xlsx')

activeSheet1 = wb1.active
maxRow = activeSheet1.max_row

activeSheet1.columns[0]
data = []           # Store row numbers
baselineData = {}

for cellObj in activeSheet1.columns[0]:
    if(cellObj.value == "Script Name"):
         data.append(cellObj.row)
x = data[0]


for k in range(x,maxRow,1):
    val = activeSheet1.cell(row=k,column=1).value
    if(type(val) == str and val != "Script Name"):
        scriptName  = activeSheet1['A' + str(k)].value
        transName = activeSheet1['B' + str(k)].value

         # Make sure the key for this state exists.
        baselineData.setdefault(scriptName, {})
        # Make sure the key for this county in this state exists
        baselineData[scriptName].setdefault(transName, {'Avg': 0, 'Max': 0, 'nine': 0, 'nine5': 0})

        # Each row represents one census tract, so increment by one.
        baselineData[scriptName][transName]['Avg'] = float(activeSheet1.cell(row=k,column=6).value)
        #print(activeSheet1.cell(row=k,column=6).value)
        #print(activeSheet1['F' + str(k)].value)

        baselineData[scriptName][transName]['Max'] = float(activeSheet1['H' + str(k)].value)
        #print(activeSheet1.cell(row=k,column=8).value)
        #print(activeSheet1['H' + str(k)].value)

        baselineData[scriptName][transName]['nine'] = float(activeSheet1['I' + str(k)].value)
        #print(activeSheet1.cell(row=k,column=9).value)
        #print(activeSheet1['I' + str(k)].value)

        baselineData[scriptName][transName]['nine5'] = float(activeSheet1['J' + str(k)].value)
        #print(activeSheet1.cell(row=k,column=10).value)
       #print(activeSheet1['J' + str(k)].value)

print('Writing results of baseline reports...')
resultFile = open('baselineData.py', 'w')
resultFile.write('allData = ' + pprint.pformat(baselineData))
resultFile.close()
print('Done.')


# ============================================= Patch Reports ===============================================

wb1 = openpyxl.load_workbook('1021_WNF9.1_OctPatch.xlsx')

activeSheet1 = wb1.active
maxRow = activeSheet1.max_row

activeSheet1.columns[0]
data2 = []           # Store row numbers
patchData = {}

for cellObj in activeSheet1.columns[0]:
    if(cellObj.value == "Script Name"):
         data2.append(cellObj.row)
x = data2[0]


for k in range(x,maxRow,1):
    val = activeSheet1.cell(row=k,column=1).value
    if(type(val) == str and val != "Script Name"):
        scriptName  = activeSheet1['A' + str(k)].value
        transName = activeSheet1['B' + str(k)].value

         # Make sure the key for this state exists.
        patchData.setdefault(scriptName, {})
        # Make sure the key for this county in this state exists
        patchData[scriptName].setdefault(transName, {'Avg': 0, 'Max': 0, 'nine': 0, 'nine5': 0})

        # Each row represents one census tract, so increment by one.
        patchData[scriptName][transName]['Avg'] = float(activeSheet1.cell(row=k,column=6).value)
        #print(activeSheet1.cell(row=k,column=6).value)
        #print(activeSheet1['F' + str(k)].value)

        patchData[scriptName][transName]['Max'] = float(activeSheet1['H' + str(k)].value)
        #print(activeSheet1.cell(row=k,column=8).value)
        #print(activeSheet1['H' + str(k)].value)

        patchData[scriptName][transName]['nine'] = float(activeSheet1['I' + str(k)].value)
        #print(activeSheet1.cell(row=k,column=9).value)
        #print(activeSheet1['I' + str(k)].value)

        patchData[scriptName][transName]['nine5'] = float(activeSheet1['J' + str(k)].value)
        #print(activeSheet1.cell(row=k,column=10).value)
       #print(activeSheet1['J' + str(k)].value)

print('Writing results of patch reports...')
resultFile = open('patchData.py', 'w')
resultFile.write('allData = ' + pprint.pformat(patchData))
resultFile.close()
print('Done.')

#i = 0
#for key in baselineData and patchData:
#    i +=1
 #   print(i,key)

key = input("Enter a script")

#for key1 in baselineData[key].keys() and patchData[key].keys():
#  print("\n"+key1+"\n")
#  print("Baseline Reports \t Oct Patch")
#  for key2 in baselineData[key][key1].keys() and patchData[key][key1].keys():
#      print(key2+": "+str(baselineData[key][key1][key2])+"\t\t"+key2+": "+str(patchData[key][key1][key2]))

tName =[]
avgVal = []
maxVal = []
n5 = []
n = []
for key1 in baselineData[key].keys():
    tName.append(key1)
    for key2 in baselineData[key][key1]:
        if key2 == "Avg":
            avgVal.append(baselineData[key][key1][key2])
        if key2 == "Max":
            maxVal.append(baselineData[key][key1][key2])
        if key2 == "nine":
            n.append(baselineData[key][key1][key2])
        if key2 == "nine5":
            n5.append(baselineData[key][key1][key2])

print(tName)
print(avgVal)
print(maxVal)
print(n)
print(n5)


