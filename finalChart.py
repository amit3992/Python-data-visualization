__author__ = 'kulkara0'

import openpyxl
from openpyxl.styles import Font, Style
import pprint
import matplotlib
import xlsxwriter

e1 = "1021_WNF9.1_OctPatch.xlsx"
e2 = "0909_WN9.1Baselines_Report.xlsx"

# ============================================= Baseline Reports ===============================================

wb1 = openpyxl.load_workbook('0909_WN9.1Baselines_Report.xlsx')

activeSheet1 = wb1.active
maxRow = activeSheet1.max_row

activeSheet1.columns[0]
data = []           # Store row numbers
baselineData = {}

for cellObj in activeSheet1.columns[0]:
    if(cellObj.value == "Script Name"):
         data.append(cellObj.row)
x = data[0]


for k in range(x,maxRow,1):
    val = activeSheet1.cell(row=k,column=1).value

    if(type(val) == str and val != "Script Name"):
        scriptName  = activeSheet1['A' + str(k)].value
        transName = activeSheet1['B' + str(k)].value

         # Make sure the key for this state exists.
        baselineData.setdefault(scriptName, {})
        # Make sure the key for this county in this state exists
        baselineData[scriptName].setdefault(transName, {'Avg': 0, 'Max': 0, 'nine': 0, 'nine5': 0})

        # Each row represents one census tract, so increment by one.
        baselineData[scriptName][transName]['Avg'] = float(activeSheet1.cell(row=k,column=6).value)
        baselineData[scriptName][transName]['Max'] = float(activeSheet1['H' + str(k)].value)
        baselineData[scriptName][transName]['nine'] = float(activeSheet1['I' + str(k)].value)
        baselineData[scriptName][transName]['nine5'] = float(activeSheet1['J' + str(k)].value)


print('Writing results of baseline reports...')
resultFile = open('baselineData.py', 'w')
resultFile.write('baselineData = ' + pprint.pformat(baselineData))
resultFile.close()
print('Done.')


# ============================================= Patch Reports ===============================================

wb1 = openpyxl.load_workbook('1021_WNF9.1_OctPatch.xlsx')

activeSheet1 = wb1.active
maxRow = activeSheet1.max_row

activeSheet1.columns[0]
data2 = []           # Store row numbers
patchData = {}

for cellObj in activeSheet1.columns[0]:
    if(cellObj.value == "Script Name"):
         data2.append(cellObj.row)
x = data2[0]


for k in range(x,maxRow,1):
    val = activeSheet1.cell(row=k,column=1).value
    if(type(val) == str and val != "Script Name"):
        scriptName  = activeSheet1['A' + str(k)].value
        transName = activeSheet1['B' + str(k)].value

         # Make sure the key for this state exists.
        patchData.setdefault(scriptName, {})
        # Make sure the key for this county in this state exists
        patchData[scriptName].setdefault(transName, {'Avg': 0, 'Max': 0, 'nine': 0, 'nine5': 0})

        # Each row represents one census tract, so increment by one.
        patchData[scriptName][transName]['Avg'] = float(activeSheet1.cell(row=k,column=6).value)
        patchData[scriptName][transName]['Max'] = float(activeSheet1['H' + str(k)].value)
        patchData[scriptName][transName]['nine'] = float(activeSheet1['I' + str(k)].value)
        patchData[scriptName][transName]['nine5'] = float(activeSheet1['J' + str(k)].value)


print('Writing results of patch reports...')
resultFile = open('patchData.py', 'w')
resultFile.write('patchData = ' + pprint.pformat(patchData))
resultFile.close()
print('Done.')


l = len(baselineData)
scripts = []
for key in baselineData.keys():
    scripts.append(key)
# ============================================= Create charts ===============================================

workbook = xlsxwriter.Workbook('newChart.xlsx')
num = 1
sheetNumber = []
worksheet1 = workbook.add_worksheet()
fontObj1 = Font(name='Times New Roman', bold=True)
styleObj1 = Style(font=fontObj1)
for i in range(2,l+2,1):
    sheetNumber.append(i)
worksheet1.write('A1', "Sheet Number")
worksheet1.write_column('A2',sheetNumber)
worksheet1.write('C1', "Script Name")
worksheet1.write_column('C2',scripts)

for key in baselineData.keys() and patchData.keys():
    worksheet = workbook.add_worksheet()
    worksheet.write('A1',"Script Name")
    worksheet.write('C1', key)
    worksheet.write('C3',"Baseline Report")
    worksheet.write('A4',"Transaction Name")
    worksheet.write('C4',"Avg")
    worksheet.write('D4',"Max")
    worksheet.write('E4',"90%")
    worksheet.write('F4',"95%")

    worksheet.write('K3',"Patch Report")
    worksheet.write('I4',"Transaction Name")
    worksheet.write('K4',"Avg")
    worksheet.write('L4',"Max")
    worksheet.write('M4',"90%")
    worksheet.write('N4',"95%")

    worksheet.write('C30',"Average value comparison chart")
    worksheet.write('K30',"Maximum value comparison chart")
    worksheet.write('T30',"90% value comparison chart")
    worksheet.write('AB30',"95% value comparison chart")

    # Initialize for every sheet
    transName = []
    baseLineAvg =[]
    patchAvg =[]
    baseLineMaxVal = []
    patchMaxVal = []
    baseLineNinety =[]
    patchNinety =[]
    baseLineNinety5 =[]
    patchNinety5 =[]

    #Create a chart object for every sheet
    avgChart = workbook.add_chart({'type':'bar'})
    maxChart = workbook.add_chart({'type':'bar'})
    ninetyChart = workbook.add_chart({'type':'bar'})
    ninety5Chart = workbook.add_chart({'type':'bar'})
    s = workbook.sheet_name+str(workbook.sheetname_count)

    for key1 in baselineData[key].keys():
        transName.append(key1)
        for key2 in baselineData[key][key1].keys():
            worksheet.write_column('A6',transName)
            if key2 == "Avg":
                baseLineAvg.append(baselineData[key][key1][key2])

            if key2 == "Max":
                baseLineMaxVal.append(baselineData[key][key1][key2])

            if key2 == "nine":
                baseLineNinety.append(baselineData[key][key1][key2])

            if key2 == "nine5":
                baseLineNinety5.append(baselineData[key][key1][key2])


            worksheet.write_column('C6', baseLineAvg)
            worksheet.write_column('D6', baseLineMaxVal)
            worksheet.write_column('E6', baseLineNinety)
            worksheet.write_column('F6', baseLineNinety5)

    tName =[]

    for k1 in patchData[key].keys():
        tName.append(k1)
        for k2 in patchData[key][k1].keys():
            worksheet.write_column('I6',tName)
            if k2 == "Avg":
                patchAvg.append(patchData[key][k1][k2])

            if k2 == "Max":
                patchMaxVal.append(patchData[key][k1][k2])

            if k2 == "nine":
                patchNinety.append(patchData[key][k1][k2])

            if k2 == "nine5":
                patchNinety5.append(patchData[key][k1][k2])

            worksheet.write_column('K6', patchAvg)
            worksheet.write_column('L6', patchMaxVal)
            worksheet.write_column('M6', patchNinety)
            worksheet.write_column('N6', patchNinety5)

    bal = len(baseLineAvg)+6
    pal = len(patchAvg)+6

    # Patch
    avgChart.add_series({
        'name': e1,
        'values': '='+s+'!$K$6:$K$'+str(pal),
        'categories': '='+s+'!$I$6:$I$'+str(pal),
        'data_labels': {'value': True, 'num_format': '#,##0.00'},
        'line':   {'color': 'blue'},
        'marker': {'type': 'square',
               'size,': 5,
               'border': {'color': 'red'},
               'fill':   {'color': 'yellow'}
                },
    })

    # Patch
    maxChart.add_series({
        'name': e1,
        'values': '='+s+'!$L$6:$L$'+str(pal),
        'categories': '='+s+'!$I$6:$I$'+str(pal),
        'data_labels': {'value': True, 'num_format': '#,##0.00'},
        'line':   {'color': 'yellow'},
        'marker': {'type': 'square',
               'size,': 5,
               'border': {'color': 'red'},
               'fill':   {'color': 'yellow'}
                },
    })

    # Patch
    ninetyChart.add_series({
        'name': e1,
        'values': '='+s+'!$M$6:$M$'+str(pal),
        'categories': '='+s+'!$I$6:$I$'+str(pal),
        'data_labels': {'value': True, 'num_format': '#,##0.00'},
        'line':   {'color': 'green'},
        'marker': {'type': 'square',
               'size,': 5,
               'border': {'color': 'red'},
               'fill':   {'color': 'yellow'}
                },
    })

     # Patch
    ninety5Chart.add_series({
        'name': e1,
        'values': '='+s+'!$N$6:$N$'+str(pal),
        'categories': '='+s+'!$I$6:$I$'+str(pal),
        'data_labels': {'value': True, 'num_format': '#,##0.00'},
        'line':   {'color': 'red'},
        'marker': {'type': 'square',
               'size,': 5,
               'border': {'color': 'red'},
               'fill':   {'color': 'yellow'}
                },
    })

    # Baseline
    avgChart.add_series({
        'name': e2,
        'values': '='+s+'!$C$6:$C$'+str(bal),
        'categories': '='+s+'!$A$6:$A$'+str(bal),
        'line':   {'color': 'red'},
        'data_labels': {'value': True, 'num_format': '#,##0.00'},
        'marker': {'type': 'square',
               'size,': 5,
               'border': {'color': 'red'},
               'fill':   {'color': 'yellow'}
                },
    })

    # Baseline
    maxChart.add_series({
        'name': e2,
        'values': '='+s+'!$D$6:$D$'+str(bal),
        'categories': '='+s+'!$A$6:$A$'+str(bal),
        'line':   {'color': 'blue'},
        'data_labels': {'value': True, 'num_format': '#,##0.00'},
        'marker': {'type': 'square',
               'size,': 5,
               'border': {'color': 'red'},
               'fill':   {'color': 'yellow'}
                },
    })

    # Baseline
    ninetyChart.add_series({
        'name': e2,
        'values': '='+s+'!$E$6:$E$'+str(bal),
        'categories': '='+s+'!$A$6:$A$'+str(bal),
        'line':   {'color': 'orange'},
        'data_labels': {'value': True, 'num_format': '#,##0.00'},
        'marker': {'type': 'square',
               'size,': 5,
               'border': {'color': 'red'},
               'fill':   {'color': 'yellow'}
                },
    })

    # Baseline
    ninety5Chart.add_series({
        'name': e2,
        'values': '='+s+'!$F$6:$F$'+str(bal),
        'categories': '='+s+'!$A$6:$A$'+str(bal),
        'line':   {'color': 'green'},
        'data_labels': {'value': True, 'num_format': '#,##0.00'},
        'marker': {'type': 'square',
               'size,': 5,
               'border': {'color': 'red'},
               'fill':   {'color': 'yellow'}
                },
    })

    # Add a chart title and some axis labels.
    avgChart.set_title ({'name': 'Comparison of Baseline Report and OctPatch for average values'})
    avgChart.set_x_axis({'name': 'Time'})
    avgChart.set_y_axis({'name': 'Transaction Name'})

    maxChart.set_title ({'name': 'Comparison of Baseline Report and OctPatch for max values'})
    maxChart.set_x_axis({'name': 'Time'})
    maxChart.set_y_axis({'name': 'Transaction Name'})

    ninetyChart.set_title ({'name': 'Comparison of Baseline Report and OctPatch'})
    ninetyChart.set_x_axis({'name': 'Time'})
    ninetyChart.set_y_axis({'name': 'Transaction Name'})

    ninety5Chart.set_title ({'name': 'Comparison of Baseline Report and OctPatch'})
    ninety5Chart.set_x_axis({'name': 'Time'})
    ninety5Chart.set_y_axis({'name': 'Transaction Name'})

    # Set an Excel chart style.
    avgChart.set_style(2)
    maxChart.set_style(2)
    ninetyChart.set_style(2)
    ninety5Chart.set_style(2)

    # Insert the chart into the worksheet (with an offset).
    worksheet.insert_chart('A32', avgChart)
    worksheet.insert_chart('I32', maxChart)
    worksheet.insert_chart('Q32', ninetyChart)
    worksheet.insert_chart('Y32', ninety5Chart)


print("Charts created for all scripts")
workbook.close()
